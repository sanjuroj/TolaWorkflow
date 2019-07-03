# -*- coding: utf-8 -*-
import openpyxl
from django.utils.translation import ugettext
from django.http import HttpResponse

from tola_management.programadmin import get_audit_log_workbook
from workflow.models import Program
from indicators.models import Indicator

EM_DASH = u'–'


def force_unicode(value):
    """
    Some values can be of type str, unicode, or django.utils.functional.__proxy__
    Force these 3 types to just be `unicode` for exporting to XLS
    """
    if type(value) == str:
        return value.decode('utf-8')
    return unicode(value)


class ExcelRendererBase(object):
    """Set of utility functions for rendering a serialized IPTT into an Excel export"""

    TITLE_START_COLUMN = 3 # 2 rows currently hidden at start, title starts at column C
    TITLE_FONT = openpyxl.styles.Font(size=18)
    HEADER_FONT = openpyxl.styles.Font(bold=True)
    LEVEL_ROW_FONT = openpyxl.styles.Font(bold=True)
    HEADER_FILL = openpyxl.styles.PatternFill('solid', 'EEEEEE')
    LEVEL_ROW_FILL = openpyxl.styles.PatternFill('solid', 'CCCCCC')
    CENTER_ALIGN = openpyxl.styles.Alignment(horizontal='center', vertical='bottom')
    RIGHT_ALIGN = openpyxl.styles.Alignment(horizontal='right', vertical='bottom')
    LEFT_ALIGN_WRAP = openpyxl.styles.Alignment(wrap_text=True)
    INDICATOR_NAME = openpyxl.styles.NamedStyle(
        name="indicator_name",
        font=openpyxl.styles.Font(underline='single'),
        alignment=openpyxl.styles.Alignment(wrap_text=True)
    )

    def __init__(self, serializer):
        self.NAME_MAP = {
            1: ugettext('Life of Program (LoP) only'),
            2: ugettext('Midline and endline'),
            3: ugettext('Annual'),
            4: ugettext('Semi-annual'),
            5: ugettext('Tri-annual'),
            # Translators: this is the measure of time (3 months)
            6: ugettext('Quarterly'),
            7: ugettext('Monthly')
        }
        self.serializer = serializer
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def add_sheet(self, name):
        sheet = self.wb.create_sheet(name)
        return sheet

    @property
    def header_columns(self):
        headers = [
            ugettext('Program ID'),
            ugettext('Indicator ID'),
            # Translators: "No." as in abbreviation for Number
            ugettext('No.'),
            ugettext('Indicator')
        ]
        if self.serializer.level_column:
            headers += [
                ugettext('Level')
            ]
        headers += [
            ugettext('Unit of measure'),
            # Translators: this is short for "Direction of Change" as in + or -
            ugettext('Change'),
            # Translators: 'C' as in Cumulative and 'NC' as in Non Cumulative
            ugettext('C / NC'),
            u'# / %',
            ugettext('Baseline')
        ]
        return headers

    def add_headers(self, sheet):
        for row, title in enumerate([
                self.serializer.report_title,
                self.serializer.report_date_range,
                self.serializer.program_name]):
            cell = sheet.cell(row=row+1, column=self.TITLE_START_COLUMN)
            cell.value = unicode(title)
            cell.font = self.TITLE_FONT
            sheet.merge_cells(
                start_row=row+1, start_column=self.TITLE_START_COLUMN,
                end_row=row+1, end_column=len(self.header_columns)
                )
        for column, header in enumerate(self.header_columns):
            cell = sheet.cell(row=4, column=column+1)
            cell.value = unicode(header)
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.fill = self.HEADER_FILL
        period_column = len(self.header_columns) + 1
        for period in self.all_periods:
            period_column = self.add_period_header(
                sheet, period_column, period
            )

    def add_period_header(self, sheet, col, period):
        for header, row in [
                (period.header, 2),
                (period.subheader, 3)
            ]:
            if header:
                cell = sheet.cell(row=row, column=col)
                cell.value = force_unicode(header)
                cell.font = self.HEADER_FONT
                cell.alignment = self.CENTER_ALIGN
                if period.tva:
                    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
        columns = [ugettext('Target'), ugettext('Actual'), ugettext('% Met')] if period.tva else [ugettext('Actual'),]
        for col_no, col_header in enumerate(columns):
            cell = sheet.cell(row=4, column=col+col_no)
            cell.value = unicode(col_header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.RIGHT_ALIGN
        return col + (3 if period.tva else 1)

    def add_level_row(self, row, sheet, level_name):
        sheet.cell(row=row, column=1).fill = self.LEVEL_ROW_FILL
        sheet.cell(row=row, column=2).fill = self.LEVEL_ROW_FILL
        cell = sheet.cell(row=row, column=3)
        cell.value = unicode(level_name)
        cell.font = self.LEVEL_ROW_FONT
        cell.fill = self.LEVEL_ROW_FILL
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=self.column_count)

    def int_cell(self, value):
        if not value:
            return None, 'General'
        return int(value), '0'

    def float_cell(self, value):
        if not value:
            return None, 'General'
        value = round(float(value), 2)
        if value == int(value):
            return int(value), '0'
        elif value == round(value, 1):
            return round(value, 1), '0.0'
        return value, '0.00'

    def percent_cell(self, value):
        if not value:
            return None, 'General'
        value = round(float(value), 4)
        if value == round(value, 2):
            return round(value, 2), '0%'
        elif value == round(value, 3):
            return round(value, 3), '0.0%'
        return value, '0.00%'

    def percent_value_cell(self, value):
        if not value:
            return None, 'General'
        value = round(float(value), 2)
        if value == int(value):
            return value/100, '0%'
        else:
            return value/100, '0.00%'

    def str_cell(self, value):
        if not value:
            return None, 'General'
        value = unicode(value)
        return value, 'General'

    def get_met_func(self, target_attribute, actual_attribute):
        def get_percent_met(indicator):
            try:
                actual = float(getattr(indicator, actual_attribute))
                target = float(getattr(indicator, target_attribute))
            except (TypeError, ValueError) as e:
                return None
            if target == 0:
                return None
            return actual / target
        return get_percent_met

    def get_number(self, indicator):
        if indicator.number_display:
            return indicator.number_display
        return indicator.number

    def add_indicator_row(self, row, sheet, indicator):
        if indicator.unit_of_measure_type == Indicator.PERCENTAGE:
            values_func = self.percent_value_cell
        else:
            values_func = self.float_cell
        indicator_columns = [
            ('program_id', self.int_cell, self.RIGHT_ALIGN, None),
            ('id', self.int_cell, self.RIGHT_ALIGN, None),
            (self.get_number, self.str_cell, self.LEFT_ALIGN_WRAP, None),
            ('name', self.str_cell, None, self.INDICATOR_NAME)
        ]
        if self.serializer.level_column:
            indicator_columns.append(
                ('old_level', self.str_cell, None, None)
            )
        indicator_columns += [
            ('unit_of_measure', self.str_cell, None, None),
            ('get_direction_of_change', self.str_cell, self.CENTER_ALIGN, 'empty_blank'),
            ('is_cumulative_display', self.str_cell, None, None),
            ('get_unit_of_measure_type', self.str_cell, self.CENTER_ALIGN, 'empty_blank'),
            ('baseline', values_func, None, None),
            ]
        for period in self.all_periods:
            if period.tva:
                indicator_columns.append(
                    (period.target_attribute, values_func, None, None)
                )
            indicator_columns.append(
                (period.actual_attribute, values_func, None, None)
            )
            if period.tva:
                indicator_columns.append(
                    (self.get_met_func(period.target_attribute, period.actual_attribute),
                     self.percent_cell, None, None)
                )
        for column, (attribute, format_func, alignment, style) in enumerate(indicator_columns):
            cell = sheet.cell(row=row, column=column+1)
            empty_blank = False
            if style == 'empty_blank':
                style = None
                empty_blank = True
            try:
                if callable(attribute):
                    value, number_format = format_func(attribute(indicator))
                elif isinstance(attribute, str):
                    value, number_format = format_func(getattr(indicator, attribute))
                else:
                    value, number_format = None, None
            except (AttributeError, TypeError, ValueError) as e:
                cell.value = None
                cell.comment = openpyxl.comments.Comment(
                    'error {} with attribute {} on indicator pk {}'.format(
                        e, attribute, indicator.pk
                        ), 'Tola System')
            else:
                if value is None:
                    value = '' if empty_blank else EM_DASH
                    alignment = self.CENTER_ALIGN
                cell.value = value
                if alignment is not None:
                    cell.alignment = alignment
                if style is not None:
                    cell.style = style
                if number_format is not None:
                    cell.number_format = number_format

    def set_column_widths(self, sheet):
        widths = [10, 10, 17, 100]
        if self.serializer.level_column:
            widths.append(12)
        widths += [30, 8, 15, 8, 12]
        for period in self.all_periods:
            widths += [12]*3 if period.tva else [12,]
        for col_no, width in enumerate(widths):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_no + 1)].width = width
        sheet.column_dimensions['A'].hidden = True
        sheet.column_dimensions['B'].hidden = True


    def render(self):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = u'attachment; filename="{}"'.format(self.serializer.filename)
        self.wb.save(response)
        return response

    def add_level_row_data(self, sheet):
        row_offset = 5
        for level_row in self.serializer.level_rows:
            self.add_level_row(row_offset, sheet, level_row['level'].display_name)
            row_offset += 1
            for indicator in level_row['indicators']:
                self.add_indicator_row(row_offset, sheet, indicator)
                row_offset += 1
        if self.serializer.blank_level_row:
            self.add_level_row(
                row_offset,
                sheet,
                ugettext('Indicators unassigned to a results framework level')
                )
            row_offset += 1
            for indicator in self.serializer.blank_level_row:
                self.add_indicator_row(row_offset, sheet, indicator)
                row_offset += 1

    def add_indicator_row_data(self, sheet):
        row_offset = 5
        for indicator in self.serializer.indicators:
            self.add_indicator_row(row_offset, sheet, indicator)
            row_offset += 1

    def add_data(self, sheet):
        if self.serializer.level_rows:
            self.add_level_row_data(sheet)
        else:
            self.add_indicator_row_data(sheet)

    @property
    def column_count(self):
        return len(self.header_columns) + sum(
            [3 if period.tva else 1 for period in self.all_periods]
            )


class FullReportExcelRenderer(ExcelRendererBase):

    def __init__(self, serializer):
        super(FullReportExcelRenderer, self).__init__(serializer)
        for frequency in self.NAME_MAP:
            serializer.frequency = frequency
            if serializer.indicators:
                sheet = self.add_sheet(self.NAME_MAP[frequency])
                self.frequency = frequency
                self.add_headers(sheet)
                self.add_data(sheet)
                self.set_column_widths(sheet)
        sheet = self.add_sheet('Change log')
        program = Program.objects.get(pk=self.serializer.program_data['pk'])
        get_audit_log_workbook(sheet, program)

    def get_indicators(self):
        return self.serializer.indicators

    @property
    def all_periods(self):
        return self.serializer.all_periods_for_frequency



class OneSheetExcelRenderer(ExcelRendererBase):

    def __init__(self, serializer):
        super(OneSheetExcelRenderer, self).__init__(serializer)
        sheet = self.add_sheet(self.NAME_MAP[serializer.frequency])
        self.add_headers(sheet)
        self.add_data(sheet)
        self.set_column_widths(sheet)

    def get_indicators(self):
        return self.serializer.indicators

    @property
    def all_periods(self):
        return self.serializer.all_periods_for_frequency
