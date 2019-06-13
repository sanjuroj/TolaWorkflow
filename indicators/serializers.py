# -*- coding: utf-8 -*-
import openpyxl
from rest_framework import serializers
from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import ugettext
from tola.l10n_utils import l10n_date_medium

from workflow.models import Program
from indicators.models import Indicator, Level, LevelTier, PeriodicTarget
from indicators.queries import IPTTIndicator
from tola_management.programadmin import get_audit_log_workbook


class LevelSerializer(serializers.ModelSerializer):
    """
    Level serializer for Program Page
    """
    level_depth = serializers.IntegerField(source='get_level_depth', read_only=True)

    class Meta:
        model = Level
        fields = [
            'id',
            'parent',
            'name',
            'assumptions',
            'program',
            'customsort',
            'ontology',
            'level_depth'
        ]
        read_only_fields = ['level_depth', 'ontology']


class LevelTierSerializer(serializers.ModelSerializer):
    """
    Level serializer for Program Page
    """
    class Meta:
        model = LevelTier
        fields = [
            'id',
            'name',
            'tier_depth'
        ]


class IndicatorSerializerMinimal(serializers.ModelSerializer):

    class Meta:
        model = Indicator
        fields = [
            'id',
            'name',
            'level',
            'level_order'
        ]


class IndicatorSerializer(serializers.ModelSerializer):
    """
    Serializer specific to the Program Page
    """
    reporting = serializers.BooleanField()
    all_targets_defined = serializers.IntegerField()
    results_count = serializers.IntegerField()
    results_with_evidence_count = serializers.IntegerField()
    over_under = serializers.IntegerField()
    target_period_last_end_date = serializers.DateField()
    level = LevelSerializer(read_only=True)
    lop_target_active = serializers.FloatField()

    class Meta:
        model = Indicator
        fields = [
            'id',
            'name',
            'number_display',
            'number',
            'level',
            'old_level',
            'level_order',
            'unit_of_measure',
            'unit_of_measure_type',
            'baseline',
            'baseline_na',
            'lop_target_active',
            'key_performance_indicator',
            'just_created',

            # DB annotations
            #  whether indicator progress towards targets is reported
            #  (min. one target period complete, one result reported):
            'reporting',
            'all_targets_defined',  # whether all targets are defined for this indicator
            'results_count',
            'results_with_evidence_count',
            'target_period_last_end_date', # last end date of last target period, for time-aware indicators
            'over_under',  # indicator progress towards targets (1: over, 0: within 15% of target, -1: under, "None": non reporting
        ]


class ProgramSerializer(serializers.ModelSerializer):
    """
    Serializer specific to the Program Page
    """
    class Meta:
        model = Program
        fields = [
            'id',
            'pk',
            'does_it_need_additional_target_periods',
            'reporting_period_start',
            'reporting_period_end',
            'results_framework',
        ]

class IPTTProgramSerializer(ProgramSerializer):
    reporting_period_start = serializers.DateField(format=None)
    reporting_period_end = serializers.DateField(format=None)


class ExcelRendererBase(object):
    """Set of utility functions for rendering a serialized IPTT into an Excel export"""

    BLANK_CELL = u'–'
    TITLE_START_COLUMN = 3 # 2 rows currently hidden at start, title starts at column C
    TITLE_FONT = openpyxl.styles.Font(size=18)
    HEADER_FONT = openpyxl.styles.Font(bold=True)
    LEVEL_ROW_FONT = openpyxl.styles.Font(bold=True)
    HEADER_FILL = openpyxl.styles.PatternFill('solid', 'EEEEEE')
    LEVEL_ROW_FILL = openpyxl.styles.PatternFill('solid', 'CCCCCC')
    CENTER_ALIGN = openpyxl.styles.Alignment(horizontal='center', vertical='bottom')
    RIGHT_ALIGN = openpyxl.styles.Alignment(horizontal='right', vertical='bottom')
    LEFT_ALIGN_WRAP = openpyxl.styles.Alignment(wrap_text=True)
    INDICATOR_NAME = openpyxl.styles.NamedStyle(
        name="indicator_name",
        font=openpyxl.styles.Font(underline='single'),
        alignment=openpyxl.styles.Alignment(wrap_text=True)
    )

    def __init__(self, serializer):
        self.NAME_MAP = {
            1: ugettext('Life of Program (LoP) only'),
            2: ugettext('Midline and endline'),
            3: ugettext('Annual'),
            4: ugettext('Semi-annual'),
            5: ugettext('Tri-annual'),
            # Translators: this is the measure of time (3 months)
            6: ugettext('Quarterly'),
            7: ugettext('Monthly')
        }
        self.serializer = serializer
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def add_sheet(self, name):
        sheet = self.wb.create_sheet(name)
        return sheet

    @property
    def header_columns(self):
        headers = [
            ugettext('Program ID'),
            ugettext('Indicator ID'),
            # Translators: "No." as in abbreviation for Number
            ugettext('No.'),
            ugettext('Indicator')
        ]
        if self.serializer.level_column:
            headers += [
                ugettext('Level')
            ]
        headers += [
            ugettext('Unit of measure'),
            # Translators: this is short for "Direction of Change" as in + or -
            ugettext('Change'),
            # Translators: 'C' as in Cumulative and 'NC' as in Non Cumulative
            ugettext('C / NC'),
            u'# / %',
            ugettext('Baseline')
        ]
        return headers

    def add_headers(self, sheet):
        for row, title in enumerate([
                self.serializer.report_title,
                self.serializer.report_date_range,
                self.serializer.program_name]):
            cell = sheet.cell(row=row+1, column=self.TITLE_START_COLUMN)
            cell.value = unicode(title)
            cell.font = self.TITLE_FONT
            sheet.merge_cells(
                start_row=row+1, start_column=self.TITLE_START_COLUMN,
                end_row=row+1, end_column=len(self.header_columns)
                )
        for column, header in enumerate(self.header_columns):
            cell = sheet.cell(row=4, column=column+1)
            cell.value = unicode(header)
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.fill = self.HEADER_FILL
        period_column = len(self.header_columns) + 1
        for period in self.all_periods:
            period_column = self.add_period_header(
                sheet, period_column, period
            )

    def add_period_header(self, sheet, col, period):
        for header, row in [
                (period.header, 2),
                (period.subheader, 3)
            ]:
            if header:
                cell = sheet.cell(row=row, column=col)
                cell.value = unicode(header)
                cell.font = self.HEADER_FONT
                cell.alignment = self.CENTER_ALIGN
                if period.tva:
                    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
        columns = [ugettext('Target'), ugettext('Actual'), ugettext('% Met')] if period.tva else [ugettext('Actual'),]
        for col_no, col_header in enumerate(columns):
            cell = sheet.cell(row=4, column=col+col_no)
            cell.value = unicode(col_header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.RIGHT_ALIGN
        return col + (3 if period.tva else 1)

    def add_level_row(self, row, sheet, level_name):
        sheet.cell(row=row, column=1).fill = self.LEVEL_ROW_FILL
        sheet.cell(row=row, column=2).fill = self.LEVEL_ROW_FILL
        cell = sheet.cell(row=row, column=3)
        cell.value = unicode(level_name)
        cell.font = self.LEVEL_ROW_FONT
        cell.fill = self.LEVEL_ROW_FILL
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=self.column_count)

    def int_cell(self, value):
        if not value:
            return None, 'General'
        return int(value), '0'

    def float_cell(self, value):
        if not value:
            return None, 'General'
        value = round(float(value), 2)
        if value == int(value):
            return int(value), '0'
        elif value == round(value, 1):
            return round(value, 1), '0.0'
        return value, '0.00'

    def percent_cell(self, value):
        if not value:
            return None, 'General'
        value = round(float(value), 4)
        if value == round(value, 2):
            return round(value, 2), '0%'
        elif value == round(value, 3):
            return round(value, 3), '0.0%'
        return value, '0.00%'

    def percent_value_cell(self, value):
        if not value:
            return None, 'General'
        value = round(float(value), 2)
        if value == int(value):
            return value/100, '0%'
        else:
            return value/100, '0.00%'

    def str_cell(self, value):
        if not value:
            return None, 'General'
        value = unicode(value)
        return value, 'General'

    def get_met_func(self, target_attribute, actual_attribute):
        def get_percent_met(indicator):
            try:
                actual = float(getattr(indicator, actual_attribute))
                target = float(getattr(indicator, target_attribute))
            except (TypeError, ValueError) as e:
                return None
            if target == 0:
                return None
            return actual / target
        return get_percent_met

    def get_number(self, indicator):
        if indicator.number_display:
            return indicator.number_display
        return indicator.number

    def add_indicator_row(self, row, sheet, indicator):
        if indicator.unit_of_measure_type == Indicator.PERCENTAGE:
            values_func = self.percent_value_cell
        else:
            values_func = self.float_cell
        indicator_columns = [
            ('program_id', self.int_cell, self.RIGHT_ALIGN, None),
            ('id', self.int_cell, self.RIGHT_ALIGN, None),
            (self.get_number, self.str_cell, self.RIGHT_ALIGN, None),
            ('name', self.str_cell, None, self.INDICATOR_NAME)
        ]
        if self.serializer.level_column:
            indicator_columns.append(
                ('old_level', self.str_cell, None, None)
            )
        indicator_columns += [
            ('unit_of_measure', self.str_cell, None, None),
            ('get_direction_of_change', self.str_cell, self.CENTER_ALIGN, 'empty_blank'),
            ('is_cumulative_display', self.str_cell, None, None),
            ('get_unit_of_measure_type', self.str_cell, self.CENTER_ALIGN, 'empty_blank'),
            ('baseline', values_func, None, None),
            ]
        for period in self.all_periods:
            if period.tva:
                indicator_columns.append(
                    (period.target_attribute, values_func, None, None)
                )
            indicator_columns.append(
                (period.actual_attribute, values_func, None, None)
            )
            if period.tva:
                indicator_columns.append(
                    (self.get_met_func(period.target_attribute, period.actual_attribute),
                     self.percent_cell, None, None)
                )
        for column, (attribute, format_func, alignment, style) in enumerate(indicator_columns):
            cell = sheet.cell(row=row, column=column+1)
            empty_blank = False
            if style == 'empty_blank':
                style = None
                empty_blank = True
            try:
                if callable(attribute):
                    value, number_format = format_func(attribute(indicator))
                elif isinstance(attribute, str):
                    value, number_format = format_func(getattr(indicator, attribute))
                else:
                    value, number_format = None, None
            except (AttributeError, TypeError, ValueError) as e:
                cell.value = None
                cell.comment = openpyxl.comments.Comment(
                    'error {} with attribute {} on indicator pk {}'.format(
                        e, attribute, indicator.pk
                        ), 'Tola System')
            else:
                if value is None:
                    value = '' if empty_blank else self.BLANK_CELL
                    alignment = self.CENTER_ALIGN
                cell.value = value
                if alignment is not None:
                    cell.alignment = alignment
                if style is not None:
                    cell.style = style
                if number_format is not None:
                    cell.number_format = number_format

    def set_column_widths(self, sheet):
        widths = [10, 10, 15, 100]
        if self.serializer.level_column:
            widths.append(12)
        widths += [30, 8, 15, 8, 12]
        for period in self.all_periods:
            widths += [12]*3 if period.tva else [12,]
        for col_no, width in enumerate(widths):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_no + 1)].width = width
        sheet.column_dimensions['A'].hidden = True
        sheet.column_dimensions['B'].hidden = True


    def render(self):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(self.serializer.filename)
        self.wb.save(response)
        return response

    def add_level_row_data(self, sheet):
        row_offset = 5
        for level_row in self.serializer.level_rows:
            self.add_level_row(row_offset, sheet, level_row['level'].display_name)
            row_offset += 1
            for indicator in level_row['indicators']:
                self.add_indicator_row(row_offset, sheet, indicator)
                row_offset += 1
        if self.serializer.blank_level_row:
            self.add_level_row(
                row_offset,
                sheet,
                ugettext('Indicators unassigned to a results framework level')
                )
            row_offset += 1
            for indicator in self.serializer.blank_level_row:
                self.add_indicator_row(row_offset, sheet, indicator)
                row_offset += 1

    def add_indicator_row_data(self, sheet):
        row_offset = 5
        for indicator in self.serializer.indicators:
            self.add_indicator_row(row_offset, sheet, indicator)
            row_offset += 1

    def add_data(self, sheet):
        if self.serializer.level_rows:
            self.add_level_row_data(sheet)
        else:
            self.add_indicator_row_data(sheet)

    @property
    def column_count(self):
        return len(self.header_columns) + sum(
            [3 if period.tva else 1 for period in self.all_periods]
            )


class FullReportExcelRenderer(ExcelRendererBase):

    def __init__(self, serializer):
        super(FullReportExcelRenderer, self).__init__(serializer)
        for frequency in self.NAME_MAP:
            serializer.frequency = frequency
            if serializer.indicators:
                sheet = self.add_sheet(self.NAME_MAP[frequency])
                self.frequency = frequency
                self.add_headers(sheet)
                self.add_data(sheet)
                self.set_column_widths(sheet)
        sheet = self.add_sheet('Change log')
        program = Program.objects.get(pk=self.serializer.program_data['pk'])
        get_audit_log_workbook(sheet, program)

    def get_indicators(self):
        return self.serializer.indicators

    @property
    def all_periods(self):
        return self.serializer.all_periods_for_frequency



class OneSheetExcelRenderer(ExcelRendererBase):

    def __init__(self, serializer):
        super(OneSheetExcelRenderer, self).__init__(serializer)
        sheet = self.add_sheet(self.NAME_MAP[serializer.frequency])
        self.add_headers(sheet)
        self.add_data(sheet)
        self.set_column_widths(sheet)

    def get_indicators(self):
        return self.serializer.indicators

    @property
    def all_periods(self):
        return self.serializer.all_periods_for_frequency


class Period:

    @classmethod
    def lop_period(cls):
        return cls(1, {
            'name': ugettext('Life of Program') 
        }, True)

    def __init__(self, frequency, period, tva=False):
        self.period = period
        self.frequency = int(frequency)
        self.tva = tva

    @property
    def period_count(self):
        return self.period['customsort']

    @property
    def target_attribute(self):
        if self.frequency == 1:
            return 'lop_target_real'
        else:
            return 'frequency_{f}_period_{p}_target'.format(f=self.frequency, p=self.period_count)

    @property
    def actual_attribute(self):
        if self.frequency == 1:
            return 'lop_actual'
        else:
            return 'frequency_{f}_period_{p}'.format(f=self.frequency, p=self.period_count)

    @property
    def start_display(self):
        return l10n_date_medium(self.period['start'])

    @property
    def end_display(self):
        return l10n_date_medium(self.period['end'])

    @property
    def header(self):
        """LOP/MidEnd/Monthly have no header, just the label, time-aware has the period name"""
        if self.frequency in [1, 2, 7]:
            return None
        return self.period['name']

    @property
    def subheader(self):
        """ Name for Lop/MidEnd/monthly is the subheader on the excel report, label for other time-aware"""
        if self.frequency in [1, 2, 7]:
            return self.period['name']
        return self.period['label']


class IPTTFullReportMixin:
    indicator_qs = IPTTIndicator.tva
    period_column_count = 3
    frequencies = [1, 2, 3, 4, 5, 6, 7]

    @property
    def report_name(self):
        return ugettext('IPTT TvA full program report')

    def annotate_indicators(self, indicators):
        self._all_indicators = indicators
        indicators_by_frequency = {}
        for frequency in self.frequencies:
            frequency_indicators = indicators.filter(target_frequency=frequency)
            indicators_by_frequency[frequency] = frequency_indicators.with_frequency_annotations(
                frequency,
                self.program_data['reporting_period_start'],
                self.program_data['reporting_period_end']
            )
        return indicators_by_frequency


    @property
    def indicators(self):
        return self._indicators.get(self.frequency, [])

    @property
    def blank_level_row(self):
        return [
            indicator for indicator in self.indicators.filter(
                level__isnull=True).with_logframe_sorting()
            ]

    @property
    def level_rows(self):
        if not self.program_data['using_results_framework']:
            return False
        return self._level_rows.get(self.frequency, [])


    def get_period(self, frequency, period_dict):
        return Period(frequency, period_dict, tva=True)

    def get_rf_levels(self):
        if self.request.get('groupby') == '2':
            sorted_levels = sorted(
                sorted(Level.objects.filter(program_id=self.program_data['pk']),
                       key=lambda level: level.ontology),
                key=lambda level: level.get_level_depth())
        else:
            sorted_levels = []
            levels = Level.objects.filter(
                program_id=self.program_data['pk'],
                parent_id__isnull=True
            )
            for level in levels:
                sorted_levels.append(level)
                sorted_levels += level.get_children()
        level_rows = {}
        for frequency in self.frequencies:
            level_rows[frequency] = []
            for level in sorted_levels:
                level_rows[frequency].append({
                    'level': level,
                    'indicators': [
                        indicator for indicator in self._indicators.get(frequency).filter(level=level)
                        ]
                })
        return level_rows


class IPTTTVAMixin:
    indicator_qs = IPTTIndicator.tva
    period_column_count = 3

    @property
    def report_name(self):
        return ugettext('IPTT TvA report')

    def annotate_indicators(self, indicators):
        frequency = int(self.request.get('frequency'))
        indicators = indicators.filter(target_frequency=frequency)
        return indicators.with_frequency_annotations(
            frequency,
            self.program_data['reporting_period_start'],
            self.program_data['reporting_period_end']
        )

    def get_period(self, frequency, period_dict):
        return Period(frequency, period_dict, tva=True)

    def get_rf_levels(self):
        if self.request.get('groupby') == '2':
            sorted_levels = sorted(
                sorted(Level.objects.filter(program_id=self.program_data['pk']),
                       key=lambda level: level.ontology),
                key=lambda level: level.get_level_depth())
        else:
            sorted_levels = []
            levels = Level.objects.filter(
                program_id=self.program_data['pk'],
                parent_id__isnull=True
            )
            for level in levels:
                sorted_levels.append(level)
                sorted_levels += level.get_children()
        for level in sorted_levels:
            level_row = {
                'level': level,
                'indicators': [indicator for indicator in self.indicators.filter(level=level)]
            }
            if level_row['indicators'] or not self.filters:
                yield level_row


class IPTTTimeperiodsMixin:
    indicator_qs = IPTTIndicator.timeperiods
    period_column_count = 1

    @property
    def report_name(self):
        return ugettext('IPTT Actuals only report')

    def annotate_indicators(self, indicators):
        frequency = int(self.request.get('frequency'))
        return indicators.with_frequency_annotations(
            frequency,
            self.program_data['reporting_period_start'],
            self.program_data['reporting_period_end']
            )

    def get_period(self, frequency, period_dict):
        return Period(frequency, period_dict, tva=False)

    def get_rf_levels(self):
        if self.request.get('groupby') == '2':
            sorted_levels = sorted(
                sorted(Level.objects.filter(program_id=self.program_data['pk']),
                       key=lambda level: level.ontology),
                key=lambda level: level.get_level_depth())
        else:
            sorted_levels = []
            levels = Level.objects.filter(
                program_id=self.program_data['pk'],
                parent_id__isnull=True
            )
            for level in levels:
                sorted_levels.append(level)
                sorted_levels += level.get_children()
        for level in sorted_levels:
            level_row = {
                'level': level,
                'indicators': [indicator for indicator in self.indicators.filter(level=level)]
            }
            if level_row['indicators'] or not self.filters:
                yield level_row

class IPTTJSONMixin:
    full_report = False
    filters = False # JSON never filters - all data sent to client and filtered in React

    def get_periods(self):
        frequency = int(self.request.get('frequency'))
        return [
            self.get_period(frequency, period_dict)
            for period_dict in PeriodicTarget.generate_for_frequency(
                int(self.request.get('frequency')))(
                    self.program_data['reporting_period_start'],
                    self.program_data['reporting_period_end']
                )
            ]

    def load_indicators(self):
        return self.indicator_qs.filter(program_id=self.program_data['pk'])


class IPTTExcelMixin(object):

    @property
    def filename(self):
        report_date = l10n_date_medium(timezone.now().date())
        return u'{name} {report_date}.xlsx'.format(name=self.report_name, report_date=report_date)

    @property
    def report_date_range(self):
        if self.frequency == 1:
            return u'{} – {}'.format(self.program_data['reporting_period_start'],
                                     self.program_data['reporting_period_end'])
        periods = self.get_periods(self.frequency)
        return u'{} – {}'.format(periods[0].start_display, periods[-1].end_display)

    @property
    def all_periods_for_frequency(self):
        yield Period.lop_period()
        for period in self.get_periods(self.frequency):
            yield period

    @property
    def column_count(self):
        return 12 + (self.level_column and 1) + (
            len(self.get_periods(self.frequency)) * self.period_column_count
            )

    def get_periods(self, frequency):
        if frequency == 1:
            return []
        periods = [
            self.get_period(frequency, period_dict)
            for period_dict in PeriodicTarget.generate_for_frequency(frequency)(
                self.program_data['reporting_period_start'],
                self.program_data['reporting_period_end']
            )
        ]
        start = int(self.request.get('start', 0))
        end = self.request.get('end', None)
        if end is not None:
            end = int(end) + 1
        return periods[start:end]

    def render(self, request):
        return self.renderer_class(self).render()


class IPTTExcelFullReportMixin(IPTTExcelMixin):
    full_report = True
    renderer_class = FullReportExcelRenderer

    def load_indicators(self):
        return self.indicator_qs.filter(program_id=self.program_data['pk'])



class IPTTSingleExcelMixin(IPTTExcelMixin):
    full_report = False
    renderer_class = OneSheetExcelRenderer

    def load_indicators(self):
        indicators = self.indicator_qs.filter(program_id=self.program_data['pk'])
        filter_params = ['sites', 'types', 'sectors', 'indicators']
        filters = {param: self.request.getlist(param) for param in self.request.viewkeys() & filter_params}
        if self.request.getlist('levels'):
            if self.program_data['using_results_framework']:
                filters['levels'] = [level.pk for levels in [
                    [level] + level.get_children()
                    for level in Level.objects.filter(pk__in=self.request.getlist('levels'))
                    ] for level in levels]
            else:
                filters['levels'] = self.request.getlist('levels')
        if self.request.getlist('tiers'):
            tier_depths = [tier.tier_depth for tier in LevelTier.objects.filter(pk__in=self.request.getlist('tiers'))]
            filters['levels'] = [
                level.pk for level in Level.objects.filter(program_id=self.program_data['pk'])
                if level.get_level_depth() in tier_depths
            ]
        if not filters:
            self.filters = False
        else:
            self.filters = filters
            filters['old_levels'] = not self.program_data['using_results_framework']
            indicators = indicators.apply_filters(**filters)
        return indicators

    @property
    def blank_level_row(self):
        return [indicator for indicator in self.indicators.filter(level__isnull=True).with_logframe_sorting()]

    @property
    def level_rows(self):
        if not self.program_data['using_results_framework'] or len(self.blank_level_row) == len(self.indicators):
            return False
        return self._level_rows


class IPTTSerializer(object):
    TIMEPERIODS_JSON = 1
    TVA_JSON = 2
    TIMEPERIODS_EXCEL = 3
    TVA_EXCEL = 4
    TVA_FULL_EXCEL = 5
    REPORT_TYPES = {
        TIMEPERIODS_JSON: None,
        TVA_JSON: None,
        TIMEPERIODS_EXCEL: [IPTTTimeperiodsMixin, IPTTSingleExcelMixin],
        TVA_EXCEL: [IPTTTVAMixin, IPTTSingleExcelMixin],
        TVA_FULL_EXCEL: [IPTTFullReportMixin, IPTTExcelFullReportMixin]
    }

    def __new__(cls, report_type, request, **kwargs):
        if report_type not in cls.REPORT_TYPES or cls.REPORT_TYPES[report_type] is None:
            raise NotImplementedError("report type not supported")
        else:
            kwargs.update({'request': request})
            kwargs.update({'report_title': ugettext('Indicator Performance Tracking Report')})
            return object.__new__(type(
                'IPTTReportSerializer', tuple(cls.REPORT_TYPES[report_type] + [IPTTSerializer]), kwargs
                ))

    def __init__(self, *args, **kwargs):
        self.program_data = ProgramSerializer(
            Program.objects.get(pk=self.request.get('programId'))
            ).data
        self._indicators = self.annotate_indicators(self.load_indicators())
        if self.program_data['using_results_framework']:
            self._level_rows = self.get_rf_levels()
        if not self.full_report:
            self.frequency = int(self.request.get('frequency'))
            self.periods = self.get_periods(self.frequency)

    @property
    def program_name(self):
        return self.program_data['name']

    @property
    def level_column(self):
        return not self.program_data['using_results_framework']

    @property
    def indicators(self):
        return self._indicators