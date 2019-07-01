"""
Utility functions for formatting cells to conform to a MC theme
"""

import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Side, Border

# MC theme colors
from openpyxl.utils import coordinate_from_string, column_index_from_string

DARK_RED = '9e1b32'
TAN = 'dad6cb'
WHITE = 'ffffff'
BLACK = '000000'


def apply_title_styling(cell):
    cell.font = Font(bold=True, color=WHITE, size=13)
    cell.fill = PatternFill(fill_type='solid', start_color=DARK_RED, end_color=DARK_RED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def apply_label_styling(cell):
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(fill_type='solid', start_color=TAN, end_color=TAN)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd = Side(style='thin', color=BLACK)
    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return cell





# openpyxl fix
# below is a terrible massive hack to make borders work on merged cells
# see: https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working


def _get_border(border_style=None, color=None):
    return Border(left=Side(border_style=border_style, color=color),
                  right=Side(border_style=border_style, color=color),
                  top=Side(border_style=border_style, color=color),
                  bottom=Side(border_style=border_style, color=color), )


def _update_style(ws, cell_range):
    start_cell, end_cell = str(cell_range).split(':')
    start_coord = coordinate_from_string(start_cell)
    end_coord = coordinate_from_string(end_cell)

    start_row = start_coord[1]
    end_row = end_coord[1]

    start_col = column_index_from_string(start_coord[0])
    end_col = column_index_from_string(end_coord[0])

    border_style = ws.cell(row=start_row, column=start_col).border.left.style
    color = ws.cell(row=start_row, column=start_col).border.left.color

    cellborder = _get_border(border_style, color)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = cellborder


def update_borders(wb):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for each_range in ws.merged_cells.ranges:
            _update_style(ws, each_range)
