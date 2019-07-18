"""
Data definitions and code related to Indicator Plan table view and XLS export
"""
from itertools import groupby
from operator import itemgetter

from django.db.models import functions, IntegerField
from django.utils.translation import (
    ugettext,
    ugettext_lazy as _
)

# Column groupings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils.cell import get_column_letter

from indicators import models
from indicators.serializers import (
    IndicatorPlanIndicatorWebSerializer,
    IndicatorPlanIndicatorExcelSerializer,
    IndicatorPlanLevelWebSerializer,
    IndicatorPlanLevelExcelSerializer
)
from indicators.xls_export_utils import apply_title_styling, apply_label_styling, update_borders, DARK_RED, TAN, WHITE, BLACK

PERFORMANCE_INDICATOR = _('Performance Indicator')
TARGETS = _('Targets')
DATA_ACQUISITION = _('Data Acquisition')
ANALYSES_AND_REPORTING = _('Analyses and Reporting')

# XLS cell widths
LARGE_CELL = 40
MEDIUM_CELL = 20
MEDIUM_SMALL_CELL = 13


# fields are either an Indicator attribute name (as str), or a callable
COLUMNS = [
    {
        'name': _('Level'),
        'category': PERFORMANCE_INDICATOR,
        'cell_width': 10,
        'field': 'tier_name_only',
        'screen_width': 7
    },
    {
        'name': _('No.'),
        'category': PERFORMANCE_INDICATOR,
        'field': 'results_aware_number',
        'screen_width': 3
    },
    {
        'name': _('Performance indicator'),
        'category': PERFORMANCE_INDICATOR,
        'cell_width': LARGE_CELL,
        'field': 'name',
        'screen_width': 25
    },
    {
        'name': _('Indicator source'),
        'category': PERFORMANCE_INDICATOR,
        'cell_width': MEDIUM_CELL,
        'field': 'source',
        'screen_width': 9
    },
    {
        'name': _('Indicator definition'),
        'category': PERFORMANCE_INDICATOR,
        'cell_width': LARGE_CELL,
        'field': 'definition',
        'screen_width': 25
    },
    {
        'name': _('Disaggregation'),
        'category': PERFORMANCE_INDICATOR,
        'cell_width': MEDIUM_CELL,
        'field': 'disaggregation',
        'screen_width': 10
    },

    {
        'name': _('Unit of measure'),
        'category': TARGETS,
        'cell_width': MEDIUM_CELL,
        'field': 'unit_of_measure',
        'screen_width': 7,
    },
    {
        'name': _('Direction of change'),
        'category': TARGETS,
        'cell_width': MEDIUM_SMALL_CELL,
        'field': 'get_direction_of_change_display',
        'screen_width': 5,
    },
    {
        'name': _('# / %'),
        'category': TARGETS,
        'cell_width': MEDIUM_SMALL_CELL,
        'field': 'get_unit_of_measure_type_display',
        'screen_width': 5,
    },
    {
        'name': _('Calculation'),
        'category': TARGETS,
        'field': 'is_cumulative',
        'screen_width': 7,
    },
    {
        'name': _('Baseline value'),
        'category': TARGETS,
        'field': 'baseline',
        'screen_width': 5,
    },
    {
        'name': _('LOP target'),
        'category': TARGETS,
        'field': 'lop_target',
        'screen_width': 5,
    },
    {
        'name': _('Rationale for target'),
        'category': TARGETS,
        'cell_width': MEDIUM_CELL,
        'field': 'rationale_for_target',
        'screen_width': 15,
    },
    {
        'name': _('Target frequency'),
        'category': TARGETS,
        'cell_width': MEDIUM_CELL,
        'field': 'get_target_frequency_display',
        'screen_width': 10,
    },

    {
        'name': _('Means of verification'),
        'category': DATA_ACQUISITION,
        'cell_width': MEDIUM_CELL,
        'field': 'means_of_verification',
        'screen_width': 10,
    },
    {
        'name': _('Data collection method'),
        'category': DATA_ACQUISITION,
        'cell_width': MEDIUM_CELL,
        'field': 'data_collection_method',
        'screen_width': 10,
    },
    {
        'name': _('Frequency of data collection'),
        'category': DATA_ACQUISITION,
        'cell_width': MEDIUM_CELL,
        'field': 'data_collection_frequency',
        'screen_width': 10,
    },
    {
        'name': _('Data points'),
        'category': DATA_ACQUISITION,
        'cell_width': MEDIUM_CELL,
        'field': 'data_points',
        'screen_width': 10,
    },
    {
        'name': _('Responsible person(s) & team'),
        'category': DATA_ACQUISITION,
        'cell_width': MEDIUM_CELL,
        'field': 'responsible_person',
        'screen_width': 10,
    },

    {
        'name': _('Method of analysis'),
        'category': ANALYSES_AND_REPORTING,
        'cell_width': MEDIUM_CELL,
        'field': 'method_of_analysis',
        'screen_width': 10,
    },
    {
        'name': _('Information use'),
        'category': ANALYSES_AND_REPORTING,
        'cell_width': MEDIUM_CELL,
        'field': 'information_use',
        'screen_width': 10,
    },
    {
        'name': _('Frequency of reporting'),
        'category': ANALYSES_AND_REPORTING,
        'cell_width': MEDIUM_CELL,
        'field': 'reporting_frequency',
        'screen_width': 10,
    },
    {
        'name': _('Quality assurance measures'),
        'category': ANALYSES_AND_REPORTING,
        'cell_width': MEDIUM_CELL,
        'field': 'quality_assurance',
        'screen_width': 10,
    },
    {
        'name': _('Data issues'),
        'category': ANALYSES_AND_REPORTING,
        'cell_width': MEDIUM_CELL,
        'field': 'data_issues',
        'screen_width': 10,
    },
    {
        'name': _('Comments'),
        'category': ANALYSES_AND_REPORTING,
        'cell_width': MEDIUM_CELL,
        'field': 'comments',
        'screen_width': 10,
    },
]


def row(indicator):
    """
    Return a row of the plan for an indicator with display values
    """
    r = []
    for col in COLUMNS:
        r.append(indicator.get(col.get('field')))
    return r


def column_names():
    """
    A list of column names
    """
    return [c['name'] for c in COLUMNS]


def column_widths(table_width):
    """
    Relative widths for sizing the html table columns
    """
    total = float(sum([col.get('screen_width') for col in COLUMNS]))
    return [int(float(col.get('screen_width'))/total*table_width) for col in COLUMNS]


def non_rf_indicator_queryset(program_id):
    """
    A QS of indicators to create the indicator plan from
    """
    return models.Indicator.objects.filter(
        program_id=program_id
    ).select_related().with_logframe_sorting().order_by(
            'old_level_pk',
            'logsort_type',
            functions.Cast('logsort_a', IntegerField()),
            functions.Cast('logsort_b', IntegerField()),
            functions.Cast('logsort_c', IntegerField()),
    )


def tier_sorted_indicator_queryset(program_id):
    """RF program with the "sort by level" option picked queryset"""
    return sorted(
        models.Level.objects.filter(program_id=program_id),
        key=lambda l: l.get_level_depth()
        )


def chain_sorted_indicator_queryset(program_id):
    """RF program with the "sort by outcome chain" option (default) picked queryset"""
    levels = []
    top_tier = models.Level.objects.filter(program_id=program_id, parent__isnull=True)
    levels += list(top_tier)
    for level in top_tier:
        levels += level.get_children()
    return levels


def get_rf_rows(level_qs, program_id):
    rows = []
    for level in [IndicatorPlanLevelWebSerializer(level_obj).data for level_obj in level_qs]:
        indicators = sorted(level.get('indicator_set'), key=itemgetter('level_order'))
        # this check is being bypassed because we are testing the IP with all rows shown
        #  - if it turns out we only want levels shown that HAVE indicators, remove the 'or true'
        #serialized = IndicatorPlanLevelWebSerializer(level)
        if indicators or True:
            rows.append(
                {
                    'row_type': 'level',
                    'row_data': level.get('display_name')
                }
            )
            rows += [{'row_type': 'indicator', 'row_data': row(i)} for i in indicators]
    unassigned_indicators = models.Indicator.objects.filter(
        program_id=program_id, level_id__isnull=True
        ).with_logframe_sorting()
    if unassigned_indicators:
        rows.append(
            {
                'row_type': 'level',
                'row_data': _('Indicators unassigned to a results framework level')
            }
        )
        rows += [
            {
                'row_type': 'indicator',
                'row_data': row(IndicatorPlanIndicatorWebSerializer(i).data)
            } for i in unassigned_indicators]
    return rows


def get_non_rf_rows(indicator_qs):
    return [{
        'row_type': 'indicator',
        'row_data': row(IndicatorPlanIndicatorWebSerializer(i).data)
        } for i in indicator_qs]


def columns_by_category():
    """
    Returns an iterable of column objects by category name
    :return:
    """
    return groupby(COLUMNS, lambda c: c['category'])


# XLS generation

START_ROW = 2
START_COLUMN = 2


def _apply_body_styling(cell):
    cell.font = Font(size=10)
    bd = Side(style='thin', color=BLACK)
    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def _set_column_widths(ws, col_num, width):
    ws.column_dimensions[get_column_letter(col_num)].width = width


def _set_row_height(ws, row_num, height):
    ws.row_dimensions[row_num].height = height


def _get_formatted_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = ugettext('Indicator plan')

    row_num = START_ROW
    col_num = START_COLUMN

    # Set widths of columns
    _set_column_widths(ws, 1, 3)  # spacer col
    for col in COLUMNS:
        width = col.get('cell_width')
        if width:
            _set_column_widths(ws, col_num, width)
        col_num += 1

    # Title
    col_num = START_COLUMN
    cell = ws.cell(row_num, col_num, ugettext('Indicator plan').encode('utf-8'))
    apply_title_styling(cell)
    ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num + len(COLUMNS)-1)
    row_num += 1

    # Category groupings
    col_num = START_COLUMN
    for category_name, columns in columns_by_category():
        num_columns = len(list(columns))
        cell = ws.cell(row_num, col_num, category_name.encode('utf-8'))
        ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num + num_columns-1)
        apply_label_styling(cell)
        col_num += num_columns
    row_num += 1

    # Column names
    col_num = START_COLUMN
    for column_name in column_names():
        cell = ws.cell(row_num, col_num, column_name.encode('utf-8'))
        apply_label_styling(cell)
        col_num += 1
    row_num += 1
    return wb, row_num


def _style_level_row(ws, row_num):
    ws.merge_cells(start_row=row_num, start_column=START_COLUMN, end_row=row_num, end_column=len(COLUMNS) + 1)
    cell = ws.cell(row_num, START_COLUMN)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(fill_type='solid', start_color=TAN, end_color=TAN)
    bd = Side(style='thin', color=BLACK)
    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    ws.row_dimensions[row_num].height = 30


def create_rf_workbook(levels, program_id):
    """
    Take an iterable of levels and return a workbook
    """
    wb, row_num = _get_formatted_workbook()
    ws = wb.active
    col_num = START_COLUMN
    for level in [IndicatorPlanLevelExcelSerializer(level_obj).data for level_obj in levels]:
        indicators = level.get('indicator_set')
        # this check is being bypassed because we are testing the IP with all rows shown
        #  - if it turns out we only want levels shown that HAVE indicators, remove the 'or true'
        if indicators or True:
            cell = ws.cell(row_num, col_num)
            cell.value = level.get('display_name')
            _style_level_row(ws, row_num)
            row_num += 1
            for indicator in indicators:
                for i, i_data in enumerate(row(indicator)):
                    cell = ws.cell(row_num, col_num, i_data.get('value'))
                    if i == 0:
                        apply_label_styling(cell)
                    else:
                        _apply_body_styling(cell)
                    cell.number_format = i_data.get('number_format', 'General')
                    col_num += 1
                col_num = START_COLUMN
                row_num += 1
    unassigned_indicators = models.Indicator.objects.filter(
        program_id=program_id, level_id__isnull=True
        ).with_logframe_sorting()
    if unassigned_indicators:
        cell = ws.cell(row_num, col_num)
        cell.value = ugettext('Indicators unassigned to a results framework level')
        _style_level_row(ws, row_num)
        row_num += 1
        for indicator in [IndicatorPlanIndicatorExcelSerializer(i).data for i in unassigned_indicators]:
            for i, i_data in enumerate(row(indicator)):
                cell = ws.cell(row_num, col_num, i_data.get('value'))
                if i == 0:
                    apply_label_styling(cell)
                else:
                    _apply_body_styling(cell)
                cell.number_format = i_data.get('number_format', 'General')
                col_num += 1
            col_num = START_COLUMN
            row_num += 1

    # fix issue with borders applied to multi-column cells
    update_borders(wb)

    return wb



def create_non_rf_workbook(indicators):
    """
    Take an iterable of indicators and return a workbook
    """
    wb, row_num = _get_formatted_workbook()
    ws = wb.active

    # rows
    col_num = START_COLUMN
    for level, level_indicators in groupby(indicators, lambda i: i.level):
        for indicator in [IndicatorPlanIndicatorExcelSerializer(i).data for i in level_indicators]:
            for i, i_data in enumerate(row(indicator)):
                val = i_data.get('value')
                number_format = i_data.get('number_format', 'General')
                cell = ws.cell(row_num, col_num, val)
                if i == 0:
                    # first column has different styling
                    apply_label_styling(cell)
                else:
                    _apply_body_styling(cell)
                cell.number_format = number_format
                col_num += 1
            col_num = START_COLUMN
            row_num += 1

        # spacer row
        for i in range(len(COLUMNS)):
            cell = ws.cell(row_num, col_num, '')
            apply_label_styling(cell)
            _set_row_height(ws, row_num, 5)
            col_num += 1

        col_num = START_COLUMN
        row_num += 1

    # delete last space row
    ws.delete_rows(row_num-1)
    _set_row_height(ws, row_num-1, None)

    # fix issue with borders applied to multi-column cells
    update_borders(wb)

    return wb

