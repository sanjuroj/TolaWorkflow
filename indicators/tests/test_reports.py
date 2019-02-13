from datetime import datetime, date
from openpyxl import Workbook
import urllib
import unittest

from django.http import QueryDict
from django.core.urlresolvers import reverse, reverse_lazy
from django.test import Client, RequestFactory, TestCase

from collections import OrderedDict
from factories.workflow_models import ProgramFactory, UserFactory
from factories.indicators_models import IndicatorFactory
from factories.indicators_models import IndicatorFactory
from factories.workflow_models import (
    ProgramFactory,
    TolaUserFactory,
    UserFactory,
)
from indicators.models import Indicator
from indicators.views.views_reports import (
    IPTT_Mixin,
    IPTTReportQuickstartView,
    IPTT_ExcelExport
    )
from tola.test.base_classes import TestBase, ScenarioBase

from workflow.models import Program


class IPTT_MixinTests(TestCase):
    """Test private methods not specifically tested in other test cases"""

    freqs = {
        Indicator.ANNUAL: 12,
        Indicator.SEMI_ANNUAL: IPTT_Mixin.MONTHS_PER_SEMIANNUAL,
        Indicator.TRI_ANNUAL: IPTT_Mixin.MONTHS_PER_TRIANNUAL,
        Indicator.QUARTERLY: IPTT_Mixin.MONTHS_PER_QUARTER,
        Indicator.MONTHLY: IPTT_Mixin.MONTHS_PER_MONTH,
        # Indicator.LOP, Indicator.MID_END, Indicator.EVENT
    }

    def setUp(self):
        self.mixin = IPTT_Mixin()
        self.user = UserFactory(first_name="Indy", last_name="Cater", username="CI")
        self.user.set_password('password')
        self.user.save()
        self.tola_user = TolaUserFactory(user=self.user)
        self.country = self.tola_user.country
        self.program = ProgramFactory(
            funding_status='Funded', reporting_period_start='2016-04-01', reporting_period_end='2020-06-01')
        self.program.country.add(self.country)
        self.program.save()
        self.indicator = IndicatorFactory(
            program=self.program, unit_of_measure_type=Indicator.NUMBER, is_cumulative=False,
            direction_of_change=Indicator.DIRECTION_OF_CHANGE_NONE, target_frequency=Indicator.LOP)
        self.request_factory = RequestFactory()
        self.client = Client()
        self.client.login(username="CI", password='password')

    def test__get_num_months(self):
        """Do we return the right number of months per period?"""

        for freq in IPTT_MixinTests.freqs:
            num_months_in_period = self.mixin._get_num_months(freq)
            self.assertEqual(num_months_in_period, IPTT_MixinTests.freqs[freq])

    def test__get_num_periods_returns_0_for_reversed_date_range(self):
        """Do we return if end date is before start date?"""

        _get_num_periods = IPTT_Mixin._get_num_periods
        start_date = date(2020, 1, 1)
        end_date = date(2019, 1, 1)

        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.ANNUAL), 0)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.SEMI_ANNUAL), 0)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.TRI_ANNUAL), 0)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.QUARTERLY), 0)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.MONTHLY), 0)

    def test__get_period_name(self):
        """Do we return the correct period names?"""

        _get_period_name = IPTT_Mixin._get_period_name

        self.assertEqual(_get_period_name(Indicator.ANNUAL), "Year")
        self.assertEqual(_get_period_name(Indicator.SEMI_ANNUAL), "Semi-annual")
        self.assertEqual(_get_period_name(Indicator.TRI_ANNUAL), "Tri-annual")
        self.assertEqual(_get_period_name(Indicator.QUARTERLY), "Quarter")
        self.assertEqual(_get_period_name(Indicator.MONTHLY), "Month")

    def test__get_first_period(self):
        """Do we calculate the first period of a date range correctly?"""

        real_start_date = date(2016, 7, 15)
        for freq in IPTT_MixinTests.freqs:
            num_months = self.mixin._get_num_months(freq)
            _get_first_period = self.mixin._get_first_period(real_start_date, num_months)

            if freq == Indicator.ANNUAL:
                self.assertEqual(_get_first_period,
                                 date(2016, 1, 1))
            elif freq == Indicator.SEMI_ANNUAL:
                self.assertEqual(_get_first_period,
                                 date(2016, 7, 1))
            elif freq == Indicator.TRI_ANNUAL:
                self.assertEqual(_get_first_period,
                                 date(2016, 5, 1))
            elif freq == Indicator.QUARTERLY:
                self.assertEqual(_get_first_period,
                                 date(2016, 7, 1))
            elif freq == Indicator.MONTHLY:
                self.assertEqual(_get_first_period,
                                 date(2016, 7, 1))
            else:
                self.fail('Unexpected target frequency' + freq)

    def test__generate_annotations(self):
        """Do we generate queryset annotations correctly?"""

        reporttype = 'timeperiods'
        filter_start_date = date(2018, 1, 1)
        filter_end_date = date(2019, 12, 31)
        num_recents = 0
        show_all = True

        self.mixin.program = Program()
        self.mixin.program.reporting_period_start = filter_start_date
        self.mixin.program.reporting_period_end = filter_end_date

        freqs = (Indicator.LOP, Indicator.MID_END, Indicator.EVENT, Indicator.ANNUAL,
                 Indicator.SEMI_ANNUAL, Indicator.TRI_ANNUAL, Indicator.QUARTERLY,
                 Indicator.MONTHLY)

        for freq in freqs:
            (report_end_date, all_date_ranges, periods_date_ranges) = self.mixin._generate_targetperiods(
                Indicator.MONTHLY
            )
            self.assertEqual(self.mixin.program.reporting_period_end, report_end_date)
            self.assertEqual(self.mixin.program.reporting_period_end, filter_end_date)

            annotations = self.mixin._generate_annotations(periods_date_ranges, freq, reporttype)
            if freq == Indicator.LOP:
                self.assertEqual(annotations, {})
            else:
                self.assertNotEqual(annotations, {})

    def test__get_num_periods(self):
        """Do we return the correct number of periods"""
        _get_num_periods = IPTT_Mixin._get_num_periods
        start_date = date(2016, 1, 15)
        end_date = date(2017, 12, 16)

        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.ANNUAL), 2)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.SEMI_ANNUAL), 4)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.TRI_ANNUAL), 6)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.QUARTERLY), 8)
        self.assertEqual(_get_num_periods(start_date, end_date, Indicator.MONTHLY), 24)

    @unittest.skip('outdated test')
    def test__generate_targetperiods(self):
        """Can we generate target periods correctly"""
        freqs = (Indicator.LOP, Indicator.MID_END, Indicator.EVENT, Indicator.ANNUAL,
                 Indicator.SEMI_ANNUAL, Indicator.TRI_ANNUAL, Indicator.QUARTERLY,
                 Indicator.MONTHLY)

        self.mixin.filter_form_initial_data = {
            'timeframe': 1,
            'numrecentperiods': 0,
            'period_start': '2018-01-01',
            'period_end': '2019-12-31'
        }
        filter_start_date = date(2018, 1, 1)
        filter_end_date = date(2019, 12, 31)
        num_recents = 0
        show_all = True
        self.mixin.program = Program()
        self.mixin.program.reporting_period_start = filter_start_date
        self.mixin.program.reporting_period_end = filter_end_date

        for freq in freqs:
            report_end_date, all_date_ranges, targetperiods = self.mixin._generate_targetperiods(freq)
            self.assertEqual(filter_end_date, report_end_date)
            self.assertEqual(len(all_date_ranges), 0)
            self.assertEqual(len(targetperiods), 0)

    @unittest.skip('outdated test')
    def test__generate_timeperiods(self):
        """Can we generate time periods correctly?"""

        freqs = (Indicator.LOP, Indicator.MID_END, Indicator.EVENT, Indicator.ANNUAL,
                 Indicator.SEMI_ANNUAL, Indicator.TRI_ANNUAL, Indicator.QUARTERLY,
                 Indicator.MONTHLY)
        filter_start_date = date(2018, 1, 1)
        filter_end_date = date(2019, 12, 31)
        num_recents = 0
        self.mixin.filter_form_initial_data = {
            'timeframe': 1,
            'numrecentperiods': num_recents,
            'period_start': '2018-01-01',
            'period_end': '2019-12-31'
        }
        self.mixin.program = Program()
        self.mixin.program.reporting_period_start = filter_start_date
        self.mixin.program.reporting_period_end = filter_end_date

        for freq in freqs:
            report_end_date, all_date_ranges, timeperiods = self.mixin._generate_targetperiods(Indicator.MONTHLY)
            self.assertEqual(report_end_date, filter_end_date, 'End dates don\'t match')
            if freq == Indicator.LOP or freq == Indicator.MID_END or freq == Indicator.EVENT:
                self.assertEqual(len(all_date_ranges), 0)
            elif freq == Indicator.ANNUAL:
                self.assertEqual(len(all_date_ranges), 2,
                                 'Unexpected number of date ranges for {0}: {1}'.format(freq, len(all_date_ranges)))
                self.assertEqual(len(timeperiods), 2,
                                 'Unexpected number of timeperiods for {0}: {1}'.format(freq, len(timeperiods)))
            elif freq == Indicator.SEMI_ANNUAL:
                self.assertEqual(len(all_date_ranges), 4,
                                 'Unexpected number of date ranges for {0}: {1}'.format(freq, len(all_date_ranges)))
                self.assertEqual(len(timeperiods), 4,
                                 'Unexpected number of timeperiods for {0}: {1}'.format(freq, len(timeperiods)))
            elif freq == Indicator.TRI_ANNUAL:
                self.assertEqual(len(all_date_ranges), 6,
                                 'Unexpected number of date ranges for {0}: {1}'.format(freq, len(all_date_ranges)))
                self.assertEqual(len(timeperiods), 6,
                                 'Unexpected number of timeperiods for {0}: {1}'.format(freq, len(timeperiods)))
            elif freq == Indicator.QUARTERLY:
                self.assertEqual(len(all_date_ranges), 8,
                                 'Unexpected number of date ranges for {0}: {1}'.format(freq, len(all_date_ranges)))
                self.assertEqual(len(timeperiods), 8,
                                 'Unexpected number of timeperiods for {0}: {1}'.format(freq, len(timeperiods)))
            elif freq == Indicator.MONTHLY:
                self.assertEqual(len(all_date_ranges), 24,
                                 'Unexpected number of date ranges for {0}: {1}'.format(freq, len(all_date_ranges)))
                self.assertEqual(len(timeperiods), 24,
                                 'Unexpected number of timeperiods for {0}: {1}'.format(freq, len(timeperiods)))

    def test__update_filter_form_initial(self):
        """Do we populate the initial filter form properly?"""

        data = {
            'csrfmiddlewaretoken': 'lolwut',
            'program': self.program.id,
            'formprefix': IPTTReportQuickstartView.FORM_PREFIX_TARGET,
            'timeframe': Indicator.LOP,
            'targetperiods': 1,
            'numrecentperiods': 1,
        }
        query_string = urllib.urlencode(data)
        formdata = QueryDict(query_string=query_string, mutable=True)
        self.mixin._update_filter_form_initial(formdata=formdata)

        filter_form_initial_data = self.mixin.filter_form_initial_data

        self.assertEqual(len(filter_form_initial_data), 4)
        self.assertNotIn('csrfmiddlewaretokeen', filter_form_initial_data)
        self.assertNotIn('program', filter_form_initial_data)

        # Dicts should have the same key/value pairs; the method first
        # strips off program and csrfmiddlewaretoken, so do that, too.
        del (data['csrfmiddlewaretoken'])
        del (data['program'])
        for k in data.keys():
            self.assertIn(k, formdata)
            # Coercing both to str because the data arg is an int
            # and the formdata arg is a unicode value
            self.assertEqual(str(data[k]), str(formdata[k]))

    def test__get_filters_with_no_periods(self):

        data = {
            'level': 3,
            'sector': 'Conflict Management',
            # TODO: Load fixtures for level, indicators
            'ind_type': 'Custom',
            'site': self.program.country.name,
            'indicators': self.indicator.id,
        }

        filters = self.mixin._get_filters(data)
        # Assert things about the returned filters
        self.assertEqual(len(filters), len(data))
        self.assertIn(data['level'], filters['level__in'])
        self.assertIn(data['sector'], filters['sector__in'])
        self.assertIn(data['ind_type'], filters['indicator_type__in'])
        self.assertIn(data['site'], filters['result__site__in'])
        self.assertIn(data['indicators'], filters['id__in'])
        self.assertEqual(data['level'], *filters['level__in'])
        self.assertEqual(data['sector'], *filters['sector__in'])
        self.assertEqual(data['ind_type'], *filters['indicator_type__in'])
        self.assertEqual(data['site'], *filters['result__site__in'])
        self.assertEqual(data['indicators'], *filters['id__in'])

        # TODO: Is it possible to make assertions about the filtered report
        # TODO: without a GET or POST?

    def test_prepare_indicators(self):
        self.skipTest('TODO: Test not implemented')

    def test_prepare_iptt_period_dateranges(self):
        self.skipTest('TODO: Test not implemented')

    # TODO: Mock the super call that invokes a non-existent get_context_data
    # call on IPTT_Mixin.
    def test_get_context_data(self):
        '''Does get_context_data return existing data untouched
        and without inserting new data?'''
        self.skipTest('TODO: Test not implemented')


class IPTT_ExcelExportTests(TestCase):

    def setUp(self):
        self.program = ProgramFactory(
            reporting_period_start=date(2015, 1, 1),
            reporting_period_end=date(2019, 12, 31),
            name='Test Program Name'
        )
        self.view = IPTT_ExcelExport()

    def test_get_filename(self):
        pass

    @unittest.skip('TODO: Implement this')
    def test_style_range(self):
        pass

    def test_add_headers(self):
        """headers should be built into the worksheet according to context data
        context variables used:
        reporttype (timeperiods/targetperiods)
        program Program
        report_start_date (date)
        report_end_date (date)
        report_date_ranges (dict) period_name : list
            - start, end, ...
        """

        context = {
            'reporttype': 'timeperiods',
            'report_start_date': date(2015, 1, 1),
            'report_end_date': date(2015, 12, 31),
            'program': self.program,
            'report_date_ranges': [
                {
                    'customsort': '0',
                    'name': 'period_1',
                    'start': date(2015, 1, 1),
                    'end': date(2015, 6, 30),
                },
                {
                    'customsort': '1',
                    'name': 'period_2',
                    'start': date(2015, 7, 1),
                    'end': date(2015, 12, 31)
                }
            ]
        }
        wb = Workbook()
        ws = wb.active
        ws = self.view.add_headers(ws, context)
        for column, header in [
            ('A', 'Program ID'),
            ('B', 'Indicator ID'),
            ('C', 'No.'),
            ('D', 'Indicator'),
            ('E', 'Level'),
            ('F', 'Unit of measure'),
            ('G', 'Change'),
            ('H', 'C / NC'),
            ('I', '# / %'),
            ('J', 'Baseline'),
            ('K', 'Target'),
            ('L', 'Actual'),
            ('M', '% Met'),
            ('N', 'Actual'),
            ('O', 'Actual')
            ]:
            self.assertEqual(ws['{0}4'.format(column)].value, header, "{0} != {1} for col {2}".format(
                ws['{0}4'.format(column)].value, header, column))
        self.assertEqual(ws['N2'].value, 'period_1')
        self.assertEqual(ws['N3'].value, 'Jan 1, 2015 - Jun 30, 2015')
        self.assertEqual(ws['O2'].value, 'period_2')
        self.assertEqual(ws['O3'].value, 'Jul 1, 2015 - Dec 31, 2015')
        context = {
            'reporttype': 'targetperiods',
            'report_start_date': date(2015, 1, 1),
            'report_end_date': date(2015, 12, 31),
            'program': self.program,
            'report_date_ranges': [
                {
                    'customsort': '0',
                    'name': 'period_1',
                    'start': date(2015, 1, 1),
                    'end': date(2015, 6, 30)
                },
                {
                    'customsort': '1',
                    'name': 'period_2',
                    'start': date(2015, 7, 1),
                    'end': date(2015, 12, 31),
                }
            ]
        }
        wb = Workbook()
        ws = wb.active
        ws = self.view.add_headers(ws, context)
        for column, header in [
            ('A', 'Program ID'),
            ('B', 'Indicator ID'),
            ('C', 'No.'),
            ('D', 'Indicator'),
            ('E', 'Level'),
            ('F', 'Unit of measure'),
            ('G', 'Change'),
            ('H', 'C / NC'),
            ('I', '# / %'),
            ('J', 'Baseline'),
            ('K', 'Target'),
            ('L', 'Actual'),
            ('M', '% Met'),
            ('N', 'Target'),
            ('O', 'Actual'),
            ('P', '% Met'),
            ('Q', 'Target'),
            ('R', 'Actual'),
            ('S', '% Met'),
            ]:
            self.assertEqual(ws['{0}4'.format(column)].value, header, "{0} != {1} for col {2}".format(
                ws['{0}4'.format(column)].value, header, column))
        self.assertEqual(ws['N2'].value, 'period_1')
        self.assertEqual(ws['N3'].value, 'Jan 1, 2015 - Jun 30, 2015')
        self.assertEqual(ws['Q2'].value, 'period_2')
        self.assertEqual(ws['Q3'].value, 'Jul 1, 2015 - Dec 31, 2015')


    def test_add_data(self):
        """does the right ws row get generated based on context data provided?
        context keys:
        indicators (list of indicators)
            -program id
            -id
            -number
            -name
            -lastlevel
            -unit_of_measure
            -direction_of_change
            -cumulative
            -unitttype
            -baseline
            -lop_target
            -lop_actual
            -lop_percent_met
        report_date_ranges keys: (targetperiods)
            -key_period_target
            -key_actual
            -key_percent_met
        report_date_ranges keys: (timeperiods)
            -key_actual"""
        indicator = OrderedDict([
            ('id', 1),
            ('number', '2.4'),
            ('name', 'indicator name'),
            ('lastlevel', 'level'),
            ('unit_of_measure', 'bananas'),
            ('direction_of_change', '+'),
            ('cumulative', 'Cumulative'),
            ('unittype', '#'),
            ('baseline', '100'),
            ('lop_target', '100'),
            ('lop_actual', '50'),
            ('lop_percent_met', '50%'),
            ('0_period_target', '25'),
            ('0_actual', '40'),
            ('0_percent_met', '90%'),
            ('1_period_target', '30'),
            ('1_actual', '80'),
            ('1_percent_met', '60%')
        ])
        context = {
            'report_date_ranges': [{'customsort': '0'}, {'customsort': '1'}],
            'program': self.program,
            'indicators': [
                indicator
            ]
        }
        wb = Workbook()
        ws = wb.active
        context['reporttype'] = 'timeperiods'
        ws = self.view.add_data(wb, ws, context)
        self.assertEqual(str(ws['A5'].value), str(self.program.id))
        for col, key in enumerate(indicator.keys()[:-6]):
            value = ws.cell(row=5, column=col+2).value
            self.assertEqual(str(indicator[key]), str(value))
        self.assertEqual(ws.cell(row=5, column=14).value, u'40')
        self.assertEqual(ws.cell(row=5, column=15).value, u'80')
        context['reporttype'] = 'targetperiods'
        ws = self.view.add_data(wb, ws, context)
        self.assertEqual(str(ws['A5'].value), str(self.program.id))
        for col, key in enumerate(indicator.keys()):
            value = ws.cell(row=5, column=col+2).value
            self.assertEqual(str(indicator[key]), str(value))

    def test_set_column_widths(self):
        widths = [10, 10, 10, 100, 12, 40, 8, 12]
        collapseds = [True, True] + [False*6]
        wb = Workbook()
        ws = wb.active
        for x, v in enumerate(['banana']*20):
            ws.cell(row=1, column=x+1).value = v
        ws = self.view.set_column_widths(ws)
        for col, width in enumerate(widths):
            column_letter = ws.cell(row=1, column=col+1).column
            self.assertEqual(ws.column_dimensions[column_letter].width,
                             width)
        for col, collapsed in enumerate(collapseds):
            column_letter = ws.cell(row=1, column=col+1).column
            self.assertEqual(ws.column_dimensions[column_letter].hidden,
                             collapsed)

    @unittest.skip('TODO: Implement this')
    def test_get(self):
        pass


class IPTT_ReportIndicatorsWithVariedStartDateTests(TestCase):
    '''Test IPTT reports which contain indicators with varied start dates'''

    def setUp(self):
        pass

    def test_get_context_data(self):
        self.skipTest('TODO: Test not implemented')

    def test_get(self):
        self.skipTest('TODO: Test not implemented')
