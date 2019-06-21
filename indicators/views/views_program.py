from operator import itemgetter
import openpyxl
from django.contrib.auth.decorators import login_required
from django.utils.translation import ugettext
from django.http import HttpResponse
from django.shortcuts import render
from workflow.serializers import LogframeProgramSerializer
from tola_management.permissions import has_program_read_access


TITLE_FONT = openpyxl.styles.Font(size=18)
HEADER_FONT = openpyxl.styles.Font(bold=True)
HEADER_FILL = openpyxl.styles.PatternFill('solid', 'EEEEEE')
TOP_LEFT_ALIGN_WRAP = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
LEVEL_ROW_FILL = openpyxl.styles.PatternFill('solid', 'CCCCCC')
BLACK_BORDER = openpyxl.styles.Side(border_style="thin", color="000000")
BORDER_TOP = openpyxl.styles.Border(top=BLACK_BORDER)

def add_title_cell(ws, row, column, value):
    cell = ws.cell(row=row, column=column)
    cell.value = value
    cell.font = TITLE_FONT
    return cell

def add_header_cell(ws, row, column, value):
    cell = ws.cell(row=row, column=column)
    cell.value = value.upper()
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    return cell

def get_child_levels(level, levels_by_pk):
    levels = [level]
    for child_pk in level['child_levels']:
        levels += get_child_levels(levels_by_pk[child_pk], levels_by_pk)
    return levels

def clean_unicode(value):
    if value is None or value is False:
        return u''
    if type(value) == str:
        return value.encode('utf-8')
    return value

@login_required
@has_program_read_access
def logframe_view(request, program):
    """
    Logframe view
    """
    serialized_program = LogframeProgramSerializer.load(program)
    context = {
        'js_context': serialized_program.data
    }
    return render(request, 'indicators/logframe/logframe.html', context)


@login_required
@has_program_read_access
def logframe_excel_view(request, program):
    """
    Logframe Excel Export view
    """
    program = LogframeProgramSerializer.load(program).data
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(ugettext('Logframe'))
    add_title_cell(ws, 1, 1, ugettext('Logframe'))
    ws.merge_cells(
        start_row=1, end_row=1,
        start_column=1, end_column=4
    )
    add_title_cell(ws, 2, 1, clean_unicode(program['name']))
    ws.merge_cells(
        start_row=2, end_row=2,
        start_column=1, end_column=4
    )
    for col, name in enumerate([
        ugettext('Result level'),
        ugettext('Indicators'),
        ugettext('Means of verification'),
        ugettext('Assumptions')
        ]):
        add_header_cell(ws, 3, col+1, name)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = 50
    levels = program['levels']
    if request.GET.get('groupby') == "2":
        sorted_levels = sorted(levels, key=itemgetter('level_depth', 'ontology'))
    else:
        levels_by_pk = {l['pk']: l for l in levels}
        sorted_levels = []
        for level in [l for l in sorted(levels, key=itemgetter('ontology')) if l['level_depth'] == 1]:
            sorted_levels += get_child_levels(level, levels_by_pk)
        levels = sorted_levels
    row = 4
    for level in sorted_levels:
        merge_start = row
        cell = ws.cell(row=row, column=1)
        cell.value = clean_unicode(level['display_name'])
        cell.alignment = TOP_LEFT_ALIGN_WRAP
        cell.fill = LEVEL_ROW_FILL
        cell = ws.cell(row=row, column=4)
        cell.value = clean_unicode(level['assumptions'])
        cell.alignment = TOP_LEFT_ALIGN_WRAP
        for indicator in sorted(level['indicators'], key=itemgetter('level_order')):
            cell = ws.cell(row=row, column=2)
            value = ugettext('Indicator')
            if indicator['level_order_display'] or level['display_ontology']:
                value += u' {}{}'.format(level['display_ontology'], indicator['level_order_display'])
            value += u': {}'.format(clean_unicode(indicator['name']))
            cell.value = value
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            cell = ws.cell(row=row, column=3)
            cell.value = clean_unicode(indicator['means_of_verification'])
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            row += 1
        if merge_start == row:
            row += 1
        else:
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=1, end_column=1
            )
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=4, end_column=4
            )
        cell = ws.cell(row=merge_start, column=1)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=2)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=3)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=4)
        cell.border = BORDER_TOP
    if program['unassigned_indicators']:
        merge_start = row
        cell = ws.cell(row=row, column=1)
        cell.value = ugettext('Indicators unassigned to a results framework level')
        cell.alignment = TOP_LEFT_ALIGN_WRAP
        cell.fill = LEVEL_ROW_FILL
        for indicator in program['unassigned_indicators']:
            cell = ws.cell(row=row, column=2)
            cell.value = u'{}: {}'.format(ugettext('Indicator'), clean_unicode(indicator['name']))
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            cell = ws.cell(row=row, column=3)
            cell.value = clean_unicode(indicator['means_of_verification'])
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            row += 1
        if merge_start == row:
            row += 1
        else:
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=1, end_column=1
            )
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=4, end_column=4
            )
        cell = ws.cell(row=merge_start, column=1)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=2)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=3)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=4)
        cell.border = BORDER_TOP
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(
        '{} - {}.xlsx'.format(program['name'], ugettext('Logframe'))
    )
    wb.save(response)
    return response
