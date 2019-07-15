# -*- coding: utf-8 -*-
import json
import csv
from StringIO import StringIO

from collections import OrderedDict
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from rest_framework import viewsets, permissions
from rest_framework.response import Response
from rest_framework import status as httpstatus
from rest_framework.validators import UniqueValidator
from rest_framework.decorators import list_route, detail_route
from rest_framework.serializers import (
    ModelSerializer,
    Serializer,
    ModelSerializer,
    CharField,
    IntegerField,
    ValidationError,
    PrimaryKeyRelatedField,
    BooleanField,
    HiddenField,
    JSONField,
    DateTimeField
)
from django.utils.translation import ugettext as _

from openpyxl import Workbook, utils
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill

from feed.views import SmallResultsSetPagination

from workflow.models import (
    Program,
    TolaUser,
    Organization,
    Country,
    Sector,
    ProgramAccess
)

from indicators.models import (
    Indicator,
    Level)

from .models import (
    ProgramAdminAuditLog,
    ProgramAuditLog
)

from tola.util import append_GAIT_dates

from .permissions import (
    HasProgramAdminAccess
)

def get_audit_log_workbook(ws, program):

    # helper for indicator name column
    def _indicator_name(indicator):
        if indicator.results_aware_number:
            return u'{} {}: {}'.format(
                _('Indicator'),
                unicode(indicator.results_aware_number),
                unicode(indicator.name),
            )
        else:
            return u'{}: {}'.format(
                _('Indicator'),
                unicode(indicator.name),
            )

    # helper for result level column
    def _result_level(indicator):
        if indicator.leveltier_name and indicator.level_display_ontology:
            return u'{} {}'.format(
                unicode(indicator.leveltier_name),
                unicode(indicator.level_display_ontology),
            )
        elif indicator.leveltier_name:
            return unicode(indicator.leveltier_name)
        else:
            return ''

    header = [
        Cell(ws, value=_("Date and Time")),
        # Translators: Number of the indicator being shown
        Cell(ws, value=_('Result Level')),
        Cell(ws, value=_('Indicator')),
        Cell(ws, value=_('User')),
        Cell(ws, value=_('Organization')),
        # Translators: Part of change log, indicates the type of change being made to a particular piece of data
        Cell(ws, value=_('Change type')),
        # Translators: Part of change log, shows what the data looked like before the changes
        Cell(ws, value=_('Previous entry')),
        # Translators: Part of change log, shows what the data looks like after the changes
        Cell(ws, value=_('New entry')),
        # Translators: Part of change log, reason for the change as entered by the user
        Cell(ws, value=_('Rationale'))
    ]
    
    title = Cell(ws, value=_("Change log"))
    title.font = Font(size=18)
    ws.append([title,])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(header))
    subtitle = Cell(ws, value=program.name)
    subtitle.font = Font(size=18)
    ws.append([subtitle,])
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(header))


    header_font = Font(bold=True)
    header_fill = PatternFill('solid', 'EEEEEE')

    for h in header:
        h.font = header_font
        h.fill = header_fill

    ws.append(header)

    alignment = Alignment(
        horizontal='general',
        vertical='top',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )

    for row in program.audit_logs.all().order_by('-date'):
        prev_string = u''
        for entry in row.diff_list:
            if entry['name'] == 'targets':
                for k, target in entry['prev'].iteritems():
                    prev_string += unicode(target['name'])+u": "+unicode(target['value'])+u"\r\n"

            else:
                prev_string += unicode(entry['pretty_name'])+u": "+unicode(entry['prev'] if entry['prev'] else _('N/A'))+u"\r\n"

        new_string = u''
        for entry in row.diff_list:
            if entry['name'] == 'targets':
                for k, target in entry['new'].iteritems():
                    new_string += unicode(target['name'])+u": "+unicode(target['value'])+u"\r\n"

            else:
                new_string += unicode(entry['pretty_name'])+u": "+unicode(entry['new'] if entry['new'] else _('N/A'))+u"\r\n"

        xl_row = [
            Cell(ws, value=row.date),
            Cell(ws, value=unicode(_result_level(row.indicator)) if row.indicator else _('N/A')),
            Cell(ws, value=unicode(_indicator_name(row.indicator)) if row.indicator else _('N/A')),
            Cell(ws, value=unicode(row.user.name)),
            Cell(ws, value=unicode(row.organization.name)),
            Cell(ws, value=unicode(row.pretty_change_type)),
            Cell(ws, value=unicode(prev_string)),
            Cell(ws, value=unicode(new_string)),
            Cell(ws, value=unicode(row.rationale))
        ]
        for cell in xl_row:
            cell.alignment = alignment
        ws.append(xl_row)

    for rd in ws.row_dimensions:
        rd.auto_size = True

    for cd in ws.column_dimensions:
        cd.auto_size = True
    widths = [20, 12, 50, 20, 15, 20, 40, 40, 40]
    for col_no, width in enumerate(widths):
        ws.column_dimensions[utils.get_column_letter(col_no + 1)].width = width
    return ws

class Paginator(SmallResultsSetPagination):
    def get_paginated_response(self , data):
        response = Response(OrderedDict([
            ('count', self.page.paginator.count),
            ('page_count', self.page.paginator.num_pages),
            ('next', self.get_next_link()),
            ('previous', self.get_previous_link()),
            ('results', data),
        ]))
        return response

class NestedSectorSerializer(Serializer):
    def to_representation(self, sector):
        return sector.id

    def to_internal_value(self, data):
        sector = Sector.objects.get(pk=data)
        return sector


class NestedCountrySerializer(Serializer):

    def to_representation(self, country):
        return country.id

    def to_internal_value(self, data):
        country = Country.objects.get(pk=data)
        return country

class ProgramAdminSerializer(ModelSerializer):
    id = IntegerField(allow_null=True, required=False)
    name = CharField(required=True, max_length=255)
    funding_status = CharField(required=True)
    gaitid = CharField(required=False, allow_blank=True, allow_null=True)
    description = CharField(allow_null=True, allow_blank=True)
    sector = NestedSectorSerializer(required=True, many=True)
    country = NestedCountrySerializer(required=True, many=True)
    auto_number_indicators = BooleanField(required=False)
    _using_results_framework = IntegerField(required=False, allow_null=True)

    def validate_country(self, values):
        if not values:
            raise ValidationError("This field may not be blank.")
        return values

    class Meta:
        model = Program
        fields = (
            'id',
            'name',
            'funding_status',
            'gaitid',
            'description',
            'sector',
            'country',
            'auto_number_indicators',
            '_using_results_framework'
        )

    def to_representation(self, program, with_aggregates=True):
        ret = super(ProgramAdminSerializer, self).to_representation(program)
        if not with_aggregates:
            return ret
        # Some n+1 queries here. If this is slow, Fix in queryset either either with rawsql or remodel.
        program_users = (
            TolaUser.objects.filter(programs__id=program.id).select_related('organization')
            | TolaUser.objects.filter(countries__program=program.id).select_related('organization')
        ).distinct()

        organizations = set([tu.organization_id for tu in program_users if tu.organization_id])
        organization_count = len(organizations)

        ret['program_users'] = len(program_users)
        ret['organizations'] = organization_count
        ret['onlyOrganizationId'] = organizations.pop() if organization_count == 1 else None
        if ret['_using_results_framework'] == Program.RF_ALWAYS:
            ret.pop('_using_results_framework')
        return ret

    @transaction.atomic
    def create(self, validated_data):
        if '_using_results_framework' in validated_data and \
                validated_data['_using_results_framework'] is None:
            validated_data.pop('_using_results_framework')
        country = validated_data.pop('country')
        sector = validated_data.pop('sector')
        if not validated_data['gaitid']:
            validated_data.pop('gaitid')
        program = super(ProgramAdminSerializer, self).create(validated_data)
        program.country.add(*country)
        program.sector.add(*sector)
        program.save()
        ProgramAdminAuditLog.created(
            program=program,
            created_by=self.context.get('request').user.tola_user,
            entry=program.admin_logged_fields,
        )
        return program

    @transaction.atomic
    def update(self, instance, validated_data):
        previous_state = instance.admin_logged_fields

        if '_using_results_framework' in validated_data and validated_data['_using_results_framework'] is None:
            validated_data['_using_results_framework'] = instance._using_results_framework

        # default for any unmigrated program is "auto" - so if someone sets their program to "not grouping" - reset it
        # to default ("auto")
        if '_using_results_framework' in validated_data and validated_data['_using_results_framework'] == instance.NOT_MIGRATED:
            validated_data['auto_number_indicators'] = True

        original_countries = instance.country.all()
        incoming_countries = validated_data.pop('country')
        added_countries = [x for x in incoming_countries if x not in original_countries]
        removed_countries = [x for x in original_countries if x not in incoming_countries]


        original_sectors = instance.sector.all()
        incoming_sectors = validated_data.pop('sector')
        added_sectors = [x for x in incoming_sectors if x not in original_sectors]
        removed_sectors = [x for x in original_sectors if x not in incoming_sectors]

        instance.country.remove(*removed_countries)
        instance.country.add(*added_countries)
        instance.sector.remove(*removed_sectors)
        instance.sector.add(*added_sectors)

        ProgramAccess.objects.filter(program=instance, country__in=removed_countries).delete()

        updated_instance = super(ProgramAdminSerializer, self).update(instance, validated_data)
        ProgramAdminAuditLog.updated(
            program=instance,
            changed_by=self.context.get('request').user.tola_user,
            old=previous_state,
            new=instance.admin_logged_fields,
        )
        return updated_instance

class ProgramAuditLogIndicatorSerializer(ModelSerializer):
    class Meta:
        model = Indicator
        fields = (
            'name',
            'leveltier_name',
            'level_display_ontology',
            'results_aware_number',
        )

class ProgramAuditLogLevelSerializer(ModelSerializer):
    class Meta:
        model = Level
        fields = (
            'name',
            'display_ontology',
        )

class ProgramAuditLogSerializer(ModelSerializer):
    id = IntegerField(allow_null=True, required=False)
    indicator = ProgramAuditLogIndicatorSerializer()
    level = ProgramAuditLogLevelSerializer()
    user = CharField(source='user.name', read_only=True)
    organization = CharField(source='organization.name', read_only=True)
    date = DateTimeField(format="%Y-%m-%d %H:%M:%S")

    class Meta:
        model = ProgramAuditLog
        fields = (
            'id',
            'date',
            'user',
            'organization',
            'indicator',
            'change_type',
            'rationale',
            'diff_list',
            'pretty_change_type',
            'level',
        )

class ProgramAdminAuditLogSerializer(ModelSerializer):
    id = IntegerField(allow_null=True, required=False)
    admin_user = CharField(source="admin_user.name", max_length=255)
    date = DateTimeField(format="%Y-%m-%d %H:%M:%S")

    class Meta:
        model = ProgramAdminAuditLog
        fields = (
            'id',
            'date',
            'admin_user',
            'change_type',
            'diff_list',
            'pretty_change_type',
        )

class ProgramAdminViewSet(viewsets.ModelViewSet):
    serializer_class = ProgramAdminSerializer
    pagination_class = Paginator
    permission_classes = [permissions.IsAuthenticated, HasProgramAdminAccess]

    def get_queryset(self):
        auth_user = self.request.user
        params = self.request.query_params

        #we have to handle this explicitly in case there are some programs without a country
        if auth_user.is_superuser:
            queryset = Program.objects.all()
        else:
            queryset = Program.objects.all().filter(country__in=auth_user.tola_user.managed_countries)

        if not auth_user.is_superuser:
            tola_user = auth_user.tola_user
            queryset = queryset.filter(
                Q(user_access=tola_user) | Q(country__users=tola_user)
            )

        programStatus = params.get('programStatus')
        if programStatus == 'Active':
            queryset = queryset.filter(funding_status='Funded')
        elif programStatus == 'Inactive':
            queryset = queryset.exclude(funding_status='Funded')

        programParam = params.getlist('programs[]')
        if programParam:
            queryset = queryset.filter(pk__in=programParam)

        countryFilter = params.getlist('countries[]')
        if countryFilter:
            queryset = queryset.filter(country__in=countryFilter)

        sectorFilter = params.getlist('sectors[]')
        if sectorFilter:
            queryset = queryset.filter(sector__in=sectorFilter)

        usersFilter = params.getlist('users[]')
        if usersFilter:
            queryset = queryset.filter(Q(user_access__id__in=usersFilter) | Q(country__in=Country.objects.filter(users__id__in=usersFilter)))

        organizationFilter = params.getlist('organizations[]')
        if organizationFilter:
            queryset = queryset.filter(
                Q(user_access__organization__in=organizationFilter) | Q(country__users__organization__in=organizationFilter)
            )

        return queryset.distinct()

    @list_route(methods=["get"])
    def program_filter_options(self, request):
        """Provides a non paginated list of countries for the frontend filter"""
        auth_user = self.request.user
        params = self.request.query_params
        queryset = Program.objects

        if not auth_user.is_superuser:
            tola_user = auth_user.tola_user
            queryset = queryset.filter(
                Q(user_access=tola_user) | Q(country__users=tola_user)
            )

        countryFilter = params.getlist('countries[]')
        if countryFilter:
            queryset = queryset.filter(country__in=countryFilter)
        programs = [{
            'id': program.id,
            'name': program.name,
        } for program in queryset.distinct().all()]
        return Response(programs)


    @detail_route(methods=['get'])
    def history(self, request, pk=None):
        program = Program.objects.get(pk=pk)
        history = (ProgramAdminAuditLog
            .objects
            .filter(program=program)
            .select_related('admin_user')
            .order_by('-date'))
        serializer = ProgramAdminAuditLogSerializer(history, many=True)
        return Response(serializer.data)

    @list_route(methods=["post"])
    def bulk_update_status(self, request):
        ids = request.data.get("ids")
        new_funding_status = request.data.get("funding_status")
        new_funding_status = new_funding_status if new_funding_status in ["Completed", "Funded"] else None
        if new_funding_status:
            to_update = Program.objects.filter(pk__in=ids)
            to_update.update(funding_status=new_funding_status)
            updated = [{
                'id': p.pk,
                'funding_status': p.funding_status,
            } for p in to_update]
            return Response(updated)
        return Response({}, status=httpstatus.HTTP_400_BAD_REQUEST)

    @detail_route(methods=["get"])
    def audit_log(self, request, pk=None):
        program = Program.objects.get(pk=pk)

        queryset = program.audit_logs.all().order_by('-date')
        page = self.paginate_queryset(list(queryset))
        if page is not None:
            serializer = ProgramAuditLogSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @detail_route(methods=["get"])
    def export_audit_log(self, request, pk=None):
        program = Program.objects.get(pk=pk)
        workbook = Workbook()
        workbook.remove(workbook.active)
        ws = workbook.create_sheet(_('Change log'))
        get_audit_log_workbook(ws, program)
        response = HttpResponse(content_type='application/ms-excel')
        filename = u'{} Audit Log {}.xlsx'.format(program.name, timezone.now().strftime('%b %d, %Y'))
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        workbook.save(response)
        return response

    @detail_route(methods=['get'], url_path='gait/(?P<gaitid>[^/.]+)')
    def gait(self, request, pk=None, gaitid=None):
        response = {}
        if gaitid is None:
            response['unique'] = True
        else:
            programs = Program.objects.filter(gaitid=gaitid)
            if pk is not None:
                programs = programs.exclude(pk=pk)
            if programs.count() == 0:
                response['unique'] = True
            else:
                response['unique'] = False
                response['gait_link'] = ('https://gait.mercycorps.org/search.vm?Mode=edit&sort_by=g.GrantTitle'
                                         '&q=&kw=&GrantNumber=&CostCenter=&GrantID={0}&GrantMin=&SSD=&USD=&'
                                         'SED=&UED=&Emergency=').format(gaitid)
        return JsonResponse(response)

    @detail_route(methods=['put'], url_path='sync_gait_dates')
    def sync_gait_dates(self, request, pk):
        program = Program.objects.get(pk=pk)

        # TODO: do something better than strings here...
        gait_error = append_GAIT_dates(program)

        program.save()

        return JsonResponse({
            'gait_error': gait_error,
        })
